#
# Froilabo deep freezer electronic temp log decoder
# Covert raw data to excel file
#
# Coded by tcy_justice <tamcy82@gmail.com>
# Last update: 17-Oct-2018
#
#

import re
import os
import json
from openpyxl import Workbook

# handle simple configs
class dataHandler:
	cname = 'config.cfg'
	def __init__(self):
		try:
			f = open(self.cname, 'r')
			c = f.read()
			f.close()
			try:
				self.data = json.loads(c)
			except:
				self.data = {}
		except:
			self.data = {}
	def get(self, item):
		if item in self.data:
			return self.data[item]
		else:
			return False
	def set(self, key, item):
		self.data[key] = item
	def exit(self):
		tmp = json.dumps(self.data)
		f = open(self.cname, 'w')
		f.write(tmp)
		f.close()

class frawProcessor:
	# db colum
	dayColumn = {
		'01' : 'B',		'02' : 'C',		'03' : 'D',		'04' : 'E',		'05' : 'F',		'06' : 'G',
		'07' : 'H',		'08' : 'I',		'09' : 'J',		'10' : 'K',		'11' : 'L',		'12' : 'M',
		'13' : 'N',		'14' : 'O',		'15' : 'P',		'16' : 'Q',		'17' : 'R',		'18' : 'S',
		'19' : 'T',		'20' : 'U',		'21' : 'V',		'22' : 'W',		'23' : 'X',		'24' : 'Y',
		'25' : 'Z',		'26' : 'AA',	'27' : 'AB',	'28' : 'AC',	'29' : 'AD',	'30' : 'AE',
		'31' : 'AF'
	}
	validLogVer = ['D01']

	# constructor
	def __init__(self):
		pass

	# load a log file
	def loadFile(self, fname):
		try:
			f = open(fname)
			# read lines in f:
			lines = f.readlines()
			# print file header
			header = re.search('[a-zA-Z0-9. ]+', lines[0])
			if header == None:
				raise Exception('Unknown file header')
			headerText = re.search('[a-zA-Z0-9. ]+', lines[0]).group(0).strip()
			for v in self.validLogVer:
				if not v in headerText.split(' '):
					raise Exception('Unknown version')
			print('File header: ' + re.search('[a-zA-Z0-9. ]+', lines[0]).group(0).strip())
			return lines
		except Exception as e:
			print(e)
			import sys
			sys.exit(1)
		except:
			print('[Error] Failed to load file')
			import sys
			sys.exit(1)

	# decode the date
	def decodeDate(self, l):
		# read the date from the file
		splitDate = [x for x in l.strip().split(" ") if x != ""]
		date = splitDate[2].split('/') # mm/dd/yyyy
		time = splitDate[3].split(':')
		date.append(time[0])
		date.append(time[1])
		date.append(time[2])
		return date

	# extract temperature
	def decodeTempColumn(self, l):
		try:
			return re.search('-?[0-9]+(.[0-9]+)?', l).group(0) # actual temp
		except:
			print('[Error] Failed to decode records')
			print('Debug message:')
			print('# String: ' + l + ' (length: ' + str(len(l)) + ')')
			print(re.search('-?[0-9]+(.[0-9]+)?', l))
			import sys
			sys.exit(1)

	# load a db file
	def loadDBFile(self, dbf):
		# is db file exist
		if os.path.exists(dbf) == True:
			from openpyxl import load_workbook
			try :
				# db exist, open
				print('Loading database: ' + dbf)
				print('')
				return load_workbook(dbf)
			except:
				print('[Error] Database error')
				print('')
				import sys
				sys.exit(1)
		else:
			# db not exist, create
			print('Create new db: ' + dbf)
			print('')
			wb	= Workbook()
			ws	= wb.active
			ws.title = "Log"
			ws['A1'] = 'Freezer Temp Log'
			return wb

	# prepare worksheet header
	def writeDBSheetHeader(self, s, j = 60):
		i = 0
		r = 0 # row
		while i <= 86400:
			s['A' + str(r+6)] = str(int(i/3600)).zfill(2) + ':' + str(int(i/60) - int(i/3600)*60).zfill(2) + ':' + str(i - int(i/60)* 60).zfill(2)
			i = i + j
			r = r + 1
		s['A1'] = 'Freezer Temp Log'
		s['A2'] = 'Year'
		s['A3'] = 'Month'
		s['A4'] = 'RAW file'
		s['A5'] = 'Day'

	# extract all items
	def extractLogs(self, lpath, lf, db, i = 60):
		print('Loading file: ' + lf)
		lines	= self.loadFile(lpath + '\\' + lf);

		# decode date
		fDate	= self.decodeDate(lines[1])
		datetext	= fDate[1] + '/' + fDate[0] + '/' + fDate[2]
		print('Log file date: ' + datetext)

		try :
			Idx = db.sheetnames.index(fDate[0]) # month
			dbSheet = db[fDate[0]]
		except:
			dbSheet = db.create_sheet(fDate[0]) # month
			self.writeDBSheetHeader(dbSheet, i)

		# writing header
		if not 'B3' in dbSheet:
			dbSheet['B3'] = fDate[0]
		if not 'B2' in dbSheet:
			dbSheet['B2'] = fDate[2]
		dbSheet[self.dayColumn[fDate[1]]+'5'] = fDate[1]

		# fill data starts on line 8 (i.e. #7 in array)
		l = 7
		limit = [(86400 / i + l), (86400 / i)]
		k = int((int(fDate[3]) * 60 * 60 + int(fDate[4]) * 60 + int(fDate[5]))/i)
		# k = int(fDate[3]) * 60 + int(fDate[4])
		print('Reading data ...')
		# print('Filing from: ' + self.dayColumn[str(fDate[1])] + str(l)) # debug
		while l <= limit[0] and k <= limit[1]:
			if len(lines[l].strip()) < 1:
				print('[Warning] Data is corrupted (file: ' + lf + ')')
				input('Press any key to continue ...');
				ok = False
				break; 
			else:
				r = self.decodeTempColumn(lines[l])
				dbKey = self.dayColumn[str(fDate[1])] + str(k+6) # column
				dbSheet[self.dayColumn[str(fDate[1])] + '4'] =  lf
				dbSheet[dbKey] = r
				l = l + 1
				k = k + 1

# main app
class frawMain:
	def main(self):
		print('##### Freeze Temp Log Decoder #####')
		print('##### Ver 1.0 Beta #####')
		print('Initializing ...')
		print('Loading configs ...')
		print('Current directory: ' + os.getcwd())
		self.path		= os.getcwd()
		self.pathdb		= self.path + '\\db\\'
		if not os.path.isdir(self.pathdb):
			os.makedirs(self.pathdb)
		# load configs
		self.configs	= dataHandler()
		self.imported	= self.configs.get('imported')
		self.interval	= self.configs.get('interval')
		if self.imported == False : self.imported = []
		# load data handler
		self.handler	= frawProcessor()
		self.eventDispatch()

	# user call events
	def eventDispatch(self):
		if self.interval == False:
			print('')
			print('Setup:')
			interval = input('Please set a time interval (s) [60]: ')
			if len(interval) == 0:
				self.interval = 60
				self.configs.set('interval', 60)
			else:
				self.interval = int(interval)
				self.configs.set('interval', self.interval)
			print('')
		print('')
		print('Options:')
		print('[1] Auto')
		print('[2] Clear configs')
		print('')
		op = input('Please choose an action [1]: ')
		print('')
		if op == '1'  or op == '':
			self.readLogDir()
		elif op == '2':
			self.resetConfig()
		else:
			print('No action')
			print('Terminating ...')

	# loop folders
	def readLogDir(self):
		listyear = os.listdir(self.path + '\\log')
		if len(listyear) == 0:
			return
		for y in listyear:
			if re.search('^2[0-9~]+$', y) and len(y) == 4 and os.path.isdir(self.path + '\\log\\'+ y):
				dbFilePath	= self.pathdb + y + '.xlsx'
				dbObj = self.handler.loadDBFile(dbFilePath)
				dbUpdate = False
				print('Reading folder (' + y + ') ...')
				listmonth = os.listdir(self.path + '\\log\\'+ y)
				if len(listmonth) > 0:
					for m in listmonth:
						if re.search('^2[0-9-~]+$', m) and os.path.isdir(self.path + '\\log\\' + y + '\\' + m):
							print('Reading folder (' + m + ') ...')
							listLog = os.listdir(self.path + '\\log\\' + y + '\\' + m)
							if len(listLog) > 0:
								for lf in listLog:
									if re.search('^[0-9~]+(.(L|l)(O|o)(G|g))$', lf) and not lf in self.imported:
										self.handler.extractLogs(self.path + '\\log\\' + y + '\\' + m, lf, dbObj, self.interval)
										dbUpdate = True
										# update list
										self.imported.append(lf)
										# empty line
										print('')
						elif re.search('^[0-9~]+(.(L|l)(O|o)(G|g))$', m) and not m in self.imported:
							self.handler.extractLogs(self.path + '\\log\\' + y, m, dbObj, self.interval)
							dbUpdate = True
							# update list
							self.imported.append(m)
							# empty line
							print('')
				print('')
				if dbUpdate:
					print('Writing database: ' + dbFilePath)
					dbObj.save(dbFilePath)
					print('Done')
				del dbFilePath, dbObj
				# empty line
				print('')

		print('Writing configs ...')
		self.configs.set('imported', self.imported)
		self.configs.exit()
		print('Done')
		print('')

	# clear all configs
	def resetConfig(self):
		print('Clearing configs ...')
		self.configs.set('imported', [])
		self.configs.set('interval', False)
		self.configs.exit()
		self.imported = []
		self.interval = False
		print('Done')
		print('')
		self.eventDispatch()

app = frawMain()
app.main()
